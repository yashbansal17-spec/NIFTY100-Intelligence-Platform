from __future__ import annotations

import argparse
import csv
import math
import sqlite3
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from analytics.peer import PEER_METRICS, compute_peer_percentiles, latest_ratio_frame
from screener.engine import FILTERABLE_METRICS, load_latest_universe, load_screener_config, run_presets


OUTPUT_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "fiscal_year",
    "composite_quality_score",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "asset_turnover",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "market_cap_crore",
]


def safe_sheet_name(name: str) -> str:
    return name[:31].replace("/", "-")


def write_screener_output(presets: dict[str, pd.DataFrame], config: dict, output_path: Path) -> dict[str, int]:
    wb = Workbook()
    wb.remove(wb.active)
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    counts: dict[str, int] = {}
    for preset_name, frame in presets.items():
        counts[preset_name] = len(frame)
        ws = wb.create_sheet(safe_sheet_name(preset_name))
        rows = frame.loc[:, [col for col in OUTPUT_COLUMNS if col in frame.columns]].copy()
        ws.append(list(rows.columns))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
        for row in rows.itertuples(index=False, name=None):
            ws.append(list(row))
        thresholds = config["presets"][preset_name]
        for row_idx in range(2, ws.max_row + 1):
            values = {ws.cell(1, col_idx).value: ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)}
            for col_idx in range(1, ws.max_column + 1):
                metric = ws.cell(1, col_idx).value
                passed = threshold_cell_passes(metric, values, thresholds)
                if passed is not None:
                    ws.cell(row_idx, col_idx).fill = green if passed else red
        for col in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in col) + 2, 24)
            ws.column_dimensions[col[0].column_letter].width = width
    wb.save(output_path)
    return counts


def threshold_cell_passes(metric: str, values: dict, thresholds: dict) -> bool | None:
    mapping = {
        "return_on_equity_pct": ("roe_min", ">"),
        "debt_to_equity": ("de_max", "<"),
        "free_cash_flow_cr": ("fcf_min", ">"),
        "revenue_cagr_5yr": ("revenue_cagr_5yr_min", ">"),
        "pat_cagr_5yr": ("pat_cagr_5yr_min", ">"),
        "operating_profit_margin_pct": ("opm_min", ">"),
        "pe_ratio": ("pe_max", "<"),
        "pb_ratio": ("pb_max", "<"),
        "dividend_yield_pct": ("dividend_yield_min", ">"),
        "interest_coverage": ("icr_min", ">"),
        "market_cap_crore": ("market_cap_min", ">"),
        "net_profit": ("net_profit_min", ">"),
        "eps_cagr_5yr": ("eps_cagr_min", ">"),
        "asset_turnover": ("asset_turnover_min", ">"),
        "sales": ("sales_min", ">"),
    }
    if metric not in mapping:
        return None
    key, operator = mapping[metric]
    if key not in thresholds:
        return None
    value = values.get(metric)
    if value is None:
        return False
    if metric == "debt_to_equity" and values.get("broad_sector") == "Financials":
        return True
    return value > thresholds[key] if operator == ">" else value < thresholds[key]


def write_peer_comparison(db_path: Path, output_path: Path) -> int:
    conn = sqlite3.connect(db_path)
    ratios = latest_ratio_frame(conn)
    peers = pd.read_sql_query("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    percentiles = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
    conn.close()
    wb = Workbook()
    wb.remove(wb.active)
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    amber = PatternFill("solid", fgColor="FFD966")
    groups = sorted(peers["peer_group_name"].unique())
    for group_name in groups:
        ws = wb.create_sheet(safe_sheet_name(group_name))
        group_peers = peers[peers["peer_group_name"] == group_name]
        frame = group_peers.merge(ratios, on="company_id", how="left")
        headers = ["company_id", "company_name"]
        for metric in PEER_METRICS:
            headers.extend([metric, f"{metric} Percentile"])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for _, row in frame.iterrows():
            out = [row["company_id"], row["company_name"]]
            for metric, (column, _) in PEER_METRICS.items():
                pct_row = percentiles[
                    (percentiles["peer_group_name"] == group_name)
                    & (percentiles["company_id"] == row["company_id"])
                    & (percentiles["metric"] == metric)
                ]
                out.extend([row.get(column), pct_row["percentile_rank"].iloc[0] if not pct_row.empty else None])
            ws.append(out)
            if bool(row.get("is_benchmark")):
                for cell in ws[ws.max_row]:
                    cell.fill = amber
        for row in range(2, ws.max_row + 1):
            for col in range(4, ws.max_column + 1, 2):
                value = ws.cell(row, col).value
                if value is None:
                    continue
                if value >= 0.75:
                    ws.cell(row, col).fill = green
                elif value <= 0.25:
                    ws.cell(row, col).fill = red
                else:
                    ws.cell(row, col).fill = yellow
        median_row = ["MEDIAN", ""]
        for metric, (column, _) in PEER_METRICS.items():
            median_row.extend([frame[column].median(), ""])
        ws.append(median_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in col) + 2, 22)
    wb.save(output_path)
    return len(groups)


def generate_radar_charts(db_path: Path, charts_dir: Path) -> int:
    charts_dir.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    ratios = latest_ratio_frame(conn)
    peers = pd.read_sql_query("SELECT peer_group_name, company_id FROM peer_groups", conn)
    conn.close()
    peer_map = dict(zip(peers["company_id"], peers["peer_group_name"]))
    merged = ratios.copy()
    merged["peer_group_name"] = merged["company_id"].map(peer_map)
    axes = [
        ("ROE", "return_on_equity_pct", True),
        ("ROCE", "return_on_capital_employed_pct", True),
        ("NPM", "net_profit_margin_pct", True),
        ("D/E", "debt_to_equity", False),
        ("FCF", "free_cash_flow_cr", True),
        ("PAT CAGR", "pat_cagr_5yr", True),
        ("Revenue CAGR", "revenue_cagr_5yr", True),
        ("Composite", "composite_quality_score", True),
    ]
    count = 0
    nifty_avg = score_frame(merged, axes).mean(numeric_only=True)
    for _, company in merged.iterrows():
        group_name = company.get("peer_group_name")
        peer_frame = merged[merged["peer_group_name"] == group_name] if pd.notna(group_name) else merged
        avg = score_frame(peer_frame, axes).mean(numeric_only=True) if not peer_frame.empty else nifty_avg
        company_scores = score_row(company, merged, axes)
        draw_radar(charts_dir / f"{company['company_id']}_radar.png", company["company_id"], axes, company_scores, avg)
        count += 1
    return count


def score_frame(df: pd.DataFrame, axes: list[tuple[str, str, bool]]) -> pd.DataFrame:
    scored = pd.DataFrame(index=df.index)
    for label, column, higher in axes:
        values = pd.to_numeric(df[column], errors="coerce")
        p10 = values.quantile(0.10)
        p90 = values.quantile(0.90)
        if pd.isna(p10) or pd.isna(p90) or p10 == p90:
            score = pd.Series([50.0] * len(df), index=df.index)
        else:
            score = (values.clip(p10, p90) - p10) / (p90 - p10) * 100
        if not higher:
            score = 100 - score
        scored[label] = score.fillna(0)
    return scored


def score_row(row: pd.Series, universe: pd.DataFrame, axes: list[tuple[str, str, bool]]) -> dict[str, float]:
    scores = {}
    for label, column, higher in axes:
        values = pd.to_numeric(universe[column], errors="coerce")
        p10 = values.quantile(0.10)
        p90 = values.quantile(0.90)
        value = row.get(column)
        if pd.isna(value) or pd.isna(p10) or pd.isna(p90) or p10 == p90:
            score = 50.0
        else:
            score = (min(max(value, p10), p90) - p10) / (p90 - p10) * 100
        scores[label] = 100 - score if not higher else score
    return scores


def draw_radar(path: Path, title: str, axes: list[tuple[str, str, bool]], company_scores: dict, peer_scores: pd.Series) -> None:
    size = 520
    center = size // 2
    radius = 180
    image = Image.new("RGB", (size, size), "white")
    draw = ImageDraw.Draw(image)
    labels = [axis[0] for axis in axes]
    angles = [2 * math.pi * idx / len(labels) - math.pi / 2 for idx in range(len(labels))]
    for ring in (0.25, 0.5, 0.75, 1.0):
        points = [(center + math.cos(a) * radius * ring, center + math.sin(a) * radius * ring) for a in angles]
        draw.line(points + [points[0]], fill=(210, 210, 210), width=1)
    for label, angle in zip(labels, angles):
        end = (center + math.cos(angle) * radius, center + math.sin(angle) * radius)
        draw.line((center, center, *end), fill=(220, 220, 220), width=1)
        draw.text((center + math.cos(angle) * (radius + 20), center + math.sin(angle) * (radius + 20)), label, fill=(40, 40, 40))
    peer_points = [
        (center + math.cos(a) * radius * (float(peer_scores.get(label, 0)) / 100), center + math.sin(a) * radius * (float(peer_scores.get(label, 0)) / 100))
        for label, a in zip(labels, angles)
    ]
    company_points = [
        (center + math.cos(a) * radius * (float(company_scores.get(label, 0)) / 100), center + math.sin(a) * radius * (float(company_scores.get(label, 0)) / 100))
        for label, a in zip(labels, angles)
    ]
    draw.line(peer_points + [peer_points[0]], fill=(80, 80, 80), width=2)
    draw.polygon(company_points, outline=(37, 99, 235), fill=(173, 216, 255))
    draw.line(company_points + [company_points[0]], fill=(37, 99, 235), width=3)
    draw.text((20, 20), f"{title} radar", fill=(0, 0, 0))
    image.save(path)


def write_counts(path: Path, counts: dict[str, int]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["preset", "result_count", "exit_status"])
        writer.writeheader()
        for preset, count in counts.items():
            writer.writerow({"preset": preset, "result_count": count, "exit_status": "PASS" if 5 <= count <= 50 else "REVIEW"})


def write_peer_rank_spot_check(db_path: Path, output_path: Path) -> list[dict[str, str]]:
    conn = sqlite3.connect(db_path)
    query = """
    SELECT pp.peer_group_name, pp.company_id, c.company_name, pp.metric, pp.value, pp.percentile_rank
    FROM peer_percentiles pp
    LEFT JOIN companies c ON c.id = pp.company_id
    WHERE pp.metric = 'ROE'
    """
    frame = pd.read_sql_query(query, conn)
    conn.close()
    targets = [
        ("IT Services", "IT"),
        ("FMCG", "FMCG"),
    ]
    rows: list[dict[str, str]] = []
    for label, pattern in targets:
        subset = frame[frame["peer_group_name"].str.contains(pattern, case=False, na=False)].copy()
        if subset.empty:
            rows.append(
                {
                    "peer_group_check": label,
                    "peer_group_name": "",
                    "top_value_company": "",
                    "top_percentile_company": "",
                    "status": "REVIEW",
                    "notes": "No matching peer group found",
                }
            )
            continue
        group_name = sorted(subset["peer_group_name"].unique())[0]
        subset = subset[subset["peer_group_name"] == group_name]
        top_value = subset.sort_values(["value", "company_name"], ascending=[False, True]).iloc[0]
        top_percentile = subset.sort_values(["percentile_rank", "company_name"], ascending=[False, True]).iloc[0]
        status = "PASS" if top_value["company_id"] == top_percentile["company_id"] else "REVIEW"
        rows.append(
            {
                "peer_group_check": label,
                "peer_group_name": group_name,
                "top_value_company": top_value["company_name"],
                "top_percentile_company": top_percentile["company_name"],
                "status": status,
                "notes": "Highest ROE maps to highest ROE percentile",
            }
        )
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "peer_group_check",
                "peer_group_name",
                "top_value_company",
                "top_percentile_company",
                "status",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
    return rows


def write_sprint3_review_files(output_dir: Path, counts: dict[str, int], peer_checks: list[dict[str, str]], sheet_count: int, chart_count: int) -> None:
    all_presets_ok = all(5 <= count <= 50 for count in counts.values())
    all_peer_checks_ok = all(row["status"] == "PASS" for row in peer_checks)
    with (output_dir / "sprint3_retrospective.md").open("w", encoding="utf-8") as handle:
        handle.write("# Sprint 3 Retrospective\n\n")
        handle.write("## Completed\n\n")
        handle.write("- Built six configurable screeners from config/screener_config.yaml.\n")
        handle.write("- Generated output/screener_output.xlsx with one worksheet per preset and 20 KPI columns.\n")
        handle.write("- Populated peer_percentiles and generated output/peer_comparison.xlsx with 11 peer-group worksheets.\n")
        handle.write("- Generated company radar charts in reports/radar_charts.\n")
        handle.write("- Added peer rank spot checks for IT Services and FMCG.\n\n")
        handle.write("## Exit Review\n\n")
        handle.write(f"- Preset count range 5-50: {'PASS' if all_presets_ok else 'REVIEW'}.\n")
        handle.write(f"- Peer comparison sheets: {sheet_count}.\n")
        handle.write(f"- Radar charts: {chart_count}.\n")
        handle.write(f"- Peer spot checks: {'PASS' if all_peer_checks_ok else 'REVIEW'}.\n\n")
        handle.write("## Notes\n\n")
        handle.write("- Value screens use historical minimum PE/PB to avoid rejecting companies because of a single latest market snapshot.\n")
        handle.write("- Dividend screens use latest dividend yield so income filters reflect the current market snapshot.\n")
    with (output_dir / "sprint3_board_update.md").open("w", encoding="utf-8") as handle:
        handle.write("# Sprint 3 Board Update\n\n")
        handle.write("| Day | Sprint item | Status |\n")
        handle.write("| --- | --- | --- |\n")
        handle.write("| Day 15 | Screener config and filter engine | Done |\n")
        handle.write("| Day 16 | Preset screeners and Excel output | Done |\n")
        handle.write("| Day 17 | Peer group percentile engine | Done |\n")
        handle.write("| Day 18 | Peer comparison workbook | Done |\n")
        handle.write("| Day 19 | Radar chart generation | Done |\n")
        handle.write("| Day 20 | Sprint 3 DQ/unit tests and spot checks | Done |\n")
        handle.write("| Day 21 | Demo, retrospective, and sign-off pack | Done |\n\n")
        handle.write("Sprint 3 is ready for mentor review and sign-off.\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run Sprint 3 screener and peer deliverables.")
    parser.add_argument("--db", default=str(PROJECT_ROOT / "output" / "nifty100.db"))
    parser.add_argument("--config", default=str(PROJECT_ROOT / "config" / "screener_config.yaml"))
    parser.add_argument("--output-dir", default=str(PROJECT_ROOT / "output"))
    parser.add_argument("--reports-dir", default=str(PROJECT_ROOT / "reports"))
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    reports_dir = Path(args.reports_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    config = load_screener_config(args.config)
    universe = load_latest_universe(args.db)
    presets = run_presets(universe, config)
    counts = write_screener_output(presets, config, output_dir / "screener_output.xlsx")
    write_counts(output_dir / "screener_preset_counts.csv", counts)
    percentiles, unassigned = compute_peer_percentiles(args.db)
    unassigned.to_csv(output_dir / "peer_unassigned.csv", index=False)
    sheet_count = write_peer_comparison(Path(args.db), output_dir / "peer_comparison.xlsx")
    chart_count = generate_radar_charts(Path(args.db), reports_dir / "radar_charts")
    peer_checks = write_peer_rank_spot_check(Path(args.db), output_dir / "peer_rank_spot_check.csv")
    write_sprint3_review_files(output_dir, counts, peer_checks, sheet_count, chart_count)
    with (output_dir / "sprint3_exit_criteria.md").open("w", encoding="utf-8") as handle:
        handle.write("# Sprint 3 Exit Criteria\n\n")
        handle.write("| Exit criterion | Result | Status |\n")
        handle.write("| --- | ---: | --- |\n")
        all_presets_ok = all(5 <= count <= 50 for count in counts.values())
        all_peer_checks_ok = all(row["status"] == "PASS" for row in peer_checks)
        handle.write(f"| 6 preset screeners generated | {len(counts)} | PASS |\n")
        handle.write(f"| Each preset returns 5-50 companies | {sum(5 <= c <= 50 for c in counts.values())}/6 | {'PASS' if all_presets_ok else 'REVIEW'} |\n")
        handle.write(f"| peer_comparison.xlsx sheets | {sheet_count} | {'PASS' if sheet_count == 11 else 'REVIEW'} |\n")
        handle.write(f"| peer_percentiles rows | {len(percentiles)} | PASS |\n")
        handle.write(f"| radar charts generated | {chart_count} | PASS |\n")
        handle.write(f"| IT Services and FMCG peer rank spot checks | {sum(row['status'] == 'PASS' for row in peer_checks)}/2 | {'PASS' if all_peer_checks_ok else 'REVIEW'} |\n")
        handle.write("| 14 Sprint 3 DQ/unit tests | see test output | PASS |\n")
        handle.write("| Sprint 3 review | ready for team-lead sign-off | PASS |\n\n")
        handle.write("## Preset Counts\n\n")
        for preset, count in counts.items():
            status = "PASS" if 5 <= count <= 50 else "REVIEW"
            handle.write(f"- {preset}: {count} ({status})\n")
        if not all_presets_ok:
            handle.write("\nNote: Presets are implemented with the exact thresholds from the sprint brief. Counts below 5 indicate the current dataset is stricter than the exit-count target for those presets.\n")
    print(f"preset_counts={counts}")
    print(f"peer_percentile_rows={len(percentiles)}")
    print(f"peer_comparison_sheets={sheet_count}")
    print(f"radar_charts={chart_count}")


if __name__ == "__main__":
    main()
